# Need to install "pip install psycopg2-binary" for db connection!
# Also used pip install dotenv for environment file

from dotenv import load_dotenv
from flask import Flask, request, jsonify
from flask_cors import CORS
import os
import csv, io
from datetime import datetime
from openpyxl import load_workbook

import psycopg2

# Load .env file
load_dotenv()

app = Flask(__name__)

CORS(app, origins="http://localhost:5173")

conn = psycopg2.connect(
    dbname="billCo",
    user="postgres",
    password=os.getenv("DB_PASSWORD"),
    host="localhost",
    port=os.getenv("DB_PORT")
)
cursor = conn.cursor()

# 2. Create a table
cursor.execute("""
CREATE TABLE IF NOT EXISTS transacts (
    pscode VARCHAR(10),
    tscode VARCHAR(15) PRIMARY KEY,
    screenresult TEXT,
    screenvendor TEXT,
    dnumnsf INT,
    dnumlate INT,
    davgdayslate INT,
    sevicted VARCHAR(3) CHECK (sevicted IN ('No','Yes')),
    drentwrittenoff REAL,
    dnonrentwrittenoff REAL,
    damoutcollections REAL,
    srenewed VARCHAR(3) CHECK (srenewed IN ('No','Yes')),   
    srent REAL,
    sfulfilledterm VARCHAR(3) CHECK (sfulfilledterm IN ('No','Yes')),
    dincome REAL,
    dtleasefrom DATE,
    dtleaseto DATE,
    dtmovein DATE,
    dtmoveout DATE,
    dwocount INT,
    dtroomearlyout DATE,
    sEmpCompany TEXT,
    sEmpPosition TEXT,
    sflex VARCHAR(5),
    sleap TEXT,
    sprevzip VARCHAR(10),
    daypaid INT,
    spaymentsource VARCHAR(50),
    dpaysourcechange INT
);
""")
conn.commit()

def process_row(row, columns):
    # Normalize keys and convert empty strings to None
    row = {k.strip().lower(): (v.strip() if v and v.strip() != '' else None) for k, v in row.items()}

    # Skip rows without primary key
    if not row.get("tscode"):
        return None

    # Parse dates
    DATE_COLUMNS = ['dtleasefrom', 'dtleaseto', 'dtmovein','dtmoveout','dtroomearlyout']
    for col in DATE_COLUMNS:
        if row.get(col):
            try:
                row[col] = datetime.strptime(row[col], "%m/%d/%Y").date()
            except ValueError:
                row[col] = None

    return [row[col] for col in columns]

primary_key = "tscode"
DATE_COLUMNS = ['dtleasefrom', 'dtleaseto', 'dtmovein','dtmoveout','dtroomearlyout']

def parse_file(content: bytes, filetype: str):
    """
    Parses CSV or XLSX content into a cleaned CSV-like file object suitable for PostgreSQL COPY.
    Skips first 5 rows, normalizes headers, formats dates, converts empty strings to NULL.
    """
    output_csv = io.StringIO()
    writer = csv.writer(output_csv)

    if filetype.lower() == "csv":
        csv_file = io.StringIO(content.decode("utf-8", errors="replace"))

        # Skip first 5 rows (trash)
        for _ in range(5):
            next(csv_file)

        reader = csv.reader(csv_file)
        headers = [h.strip().lower() for h in next(reader)]
        
        for row in reader:
            row_dict = {k: v.strip() if v.strip() != '' else None for k, v in zip(headers, row)}
            # Parse dates
            for col in DATE_COLUMNS:
                if row_dict.get(col):
                    try:
                        row_dict[col] = datetime.strptime(row_dict[col], "%m/%d/%Y").date().isoformat()
                    except (ValueError, TypeError):
                        row_dict[col] = None
            writer.writerow([row_dict.get(col) for col in headers])
    
    elif filetype.lower() == "xlsx":
        xlsx_file = io.BytesIO(content)
        wb = load_workbook(xlsx_file, data_only=True)
        ws = wb.active

        # Skip first 5 rows
        rows = list(ws.iter_rows(values_only=True))[5:]
        headers = [str(h).strip().lower() for h in rows[0]]  # first row after skip as header

        for row in rows[1:]:
            row_dict = {k: (str(v).strip() if v is not None and str(v).strip() != '' else None)
                        for k, v in zip(headers, row)}
            # Parse dates
            for col in DATE_COLUMNS:
                if row_dict.get(col):
                    try:
                        row_dict[col] = datetime.strptime(row_dict[col], "%m/%d/%Y").date().isoformat()
                    except (ValueError, TypeError):
                        row_dict[col] = None
            writer.writerow([row_dict.get(col) for col in headers])

    else:
        raise ValueError("Unsupported file type")

    output_csv.seek(0)
    return headers, output_csv

@app.route("/upload", methods=["POST"])
def upload_file():
    if "transact" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    # file = request.files["transact"]

    # filename = file.filename.lower()
    # if filename.endswith(".csv"):
    #     filetype = "csv"
    # elif filename.endswith(".xlsx"):
    #     filetype = "xlsx"
    # else:
    #     return jsonify({"error": "Unsupported file type"}), 400

    # content = file.read()  # reads the entire file into memory as bytes

    # try:
    #     headers, clean_csv = parse_file(content, filetype)
        
    #     # Create temp table
    #     # check if temp table exists already
    #     cursor.execute("DROP TABLE IF EXISTS transacts_temp;")
    #     temp_table = "transacts_temp"
    #     columns_sql = ", ".join([f"{col} TEXT" for col in headers])
    #     cursor.execute(f"""
    #                    CREATE TEMP TABLE transacts_temp (
    #                         pscode VARCHAR(10),
    #                         tscode VARCHAR(15) PRIMARY KEY,
    #                         screenresult TEXT,
    #                         screenvendor TEXT,
    #                         dnumnsf INT,
    #                         dnumlate INT,
    #                         davgdayslate INT,
    #                         sevicted VARCHAR(3) CHECK (sevicted IN ('No','Yes')),
    #                         drentwrittenoff REAL,
    #                         dnonrentwrittenoff REAL,
    #                         damoutcollections REAL,
    #                         srenewed VARCHAR(3) CHECK (srenewed IN ('No','Yes')),   
    #                         srent REAL,
    #                         sfulfilledterm VARCHAR(3) CHECK (sfulfilledterm IN ('No','Yes')),
    #                         dincome REAL,
    #                         dtleasefrom DATE,
    #                         dtleaseto DATE,
    #                         dtmovein DATE,
    #                         dtmoveout DATE,
    #                         dwocount INT,
    #                         dtroomearlyout DATE,
    #                         sEmpCompany TEXT,
    #                         sEmpPosition TEXT,
    #                         sflex VARCHAR(5),
    #                         sleap TEXT,
    #                         sprevzip VARCHAR(10),
    #                         daypaid INT,
    #                         spaymentsource VARCHAR(50),
    #                         dpaysourcechange INT
    #                     );;""")
        
    #     cursor.execute(f"ALTER TABLE {temp_table} ADD UNIQUE ({primary_key});")
    #     # COPY into temp table
    #     cursor.copy_expert(f"COPY {temp_table} ({', '.join(headers)}) FROM STDIN WITH CSV", clean_csv)

    #     # Upsert into main table
    #     primary_key = "tscode"
    #     update_clause = ", ".join([f"{col}=EXCLUDED.{col}" for col in headers if col != primary_key])
    #     upsert_sql = f"""
    #     INSERT INTO transacts ({', '.join(headers)})
    #     SELECT *
    #     FROM {temp_table}
    #     ON CONFLICT ({primary_key}) DO UPDATE
    #     SET {update_clause};
    #     """
    #     cursor.execute(upsert_sql)
    #     conn.commit()

    #     return jsonify({"success": True, "rows": len(headers)})
    # except Exception as e:
    #     conn.rollback()
    #     return jsonify({"error": str(e)}), 500



    file = request.files["transact"]
    filename = file.filename
    content = file.read()  # bytes of the file

    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        try:
            content.decode("utf-8")
        except UnicodeDecodeError:
            return jsonify({"message": f"{filename} was detected csv file to be not UTF-8 encoded, no importing could be done"}), 400 
        batch_size = 1000  # adjust as needed
        batch = []

        primary_key = "tscode"
        DATE_COLUMNS = ['dtleasefrom', 'dtleaseto', 'dtmovein','dtmoveout','dtroomearlyout']

        # Read CSV from bytes
        csv_file = io.StringIO(content.decode("utf-8", errors="replace"))
        # Skip first 5 rows (trash)
        for _ in range(5):
            next(csv_file)

        # Now DictReader sees the 6th row as the header
        reader = csv.DictReader(csv_file)
        reader.fieldnames = [h.strip().lower() for h in reader.fieldnames]
        print(reader.fieldnames)

        # Build a template SQL string (columns are dynamic per first row)
        first_row = next(reader)
        first_row = {k.strip().lower(): (v.strip() if v.strip() != '' else None) for k, v in first_row.items()}

        columns = list(first_row.keys())
        placeholders = ', '.join(['%s'] * len(columns))
        columns_sql = ', '.join(columns)
        update_clause = ', '.join([f"{col}=EXCLUDED.{col}" for col in columns if col != primary_key])

        sql = f"""
        INSERT INTO transacts ({columns_sql})
        VALUES ({placeholders})
        ON CONFLICT ({primary_key}) DO UPDATE
        SET {update_clause}
        """

        # Add first row
        row_values = process_row(first_row, columns)
        if row_values:
            batch.append(row_values)

        # Process remaining rows
        for row in reader:
            row_values = process_row(row, columns)
            if row_values:
                batch.append(row_values)

            # Execute batch if full
            if len(batch) >= batch_size:
                cursor.executemany(sql, batch)
                batch = []

        # Insert any remaining rows
        if batch:
            cursor.executemany(sql, batch)
        # # Convert bytes to a file-like object for csv.DictReader
        # csv_file = io.StringIO(content.decode("utf-8"))
        # reader = csv.DictReader(csv_file)

        # # Convert headers to lowercase
        # reader.fieldnames = [h.lower() for h in reader.fieldnames]

        # for row in reader:
        #     # Normalize row keys and convert empty strings to None
        #     row = {k.strip().lower(): (v.strip() if v.strip() != '' else None) for k, v in row.items()}

        #     if not row.get("tscode"):
        #         continue
        #     # Parse dates
        #     DATE_COLUMNS = ['dtleasefrom', 'dtleaseto', 'dtmovein','dtmoveout','dtroomearlyout']
        #     for col in DATE_COLUMNS:
        #         if row.get(col):
        #             try:
        #                 row[col] = datetime.strptime(row[col], "%m/%d/%Y").date()
        #             except ValueError:
        #                 row[col] = None  # Invalid date → NULL

        #     columns = list(row.keys())
        #     values = [row[col] for col in columns]

        #     placeholders = ', '.join(['%s'] * len(columns))
        #     columns_sql = ', '.join(columns)

        #     # Exclude primary key from update clause
        #     update_clause = ', '.join([f"{col}=EXCLUDED.{col}" for col in columns if col != "tscode"])

        #     sql = f"""
        #     INSERT INTO transacts ({columns_sql})
        #     VALUES ({placeholders})
        #     ON CONFLICT (tscode) DO UPDATE
        #     SET {update_clause}
        #     """
        #     cursor.execute(sql, values)  
    elif ext == ".xlsx":
        xlsx_file = io.BytesIO(content)
        wb = load_workbook(xlsx_file, data_only=True)
        ws = wb.active

        # Skip the first 5 rows
        rows = ws.iter_rows(values_only=True)
        for _ in range(5):
            next(rows)

        # Now the next row is your header
        headers = [cell.lower() for cell in next(rows)]

        # Iterate through the remaining rows
        for row in rows:
            row_dict = dict(zip(headers, row))
            # Prepare columns and values for INSERT
            # Parse dates
            if not row_dict.get("tscode"): # If primary key doesn't exist, skip it
                continue
                
            DATE_COLUMNS = ['dtleasefrom', 'dtleaseto', 'dtmovein','dtmoveout','dtroomearlyout']
            for col in DATE_COLUMNS:
                if row_dict[col]:
                    try:
                        row_dict[col] = datetime.strptime(row_dict[col], "%m/%d/%Y").date()
                    except ValueError:
                        row_dict[col] = None  # Invalid date → NULL

            # columns = list(row_dict.keys())
            # values = [row_dict[col] for col in columns]

            # placeholders = ', '.join(['%s'] * len(columns))
            # columns_sql = ', '.join(columns)

            # sql = f"INSERT INTO transacts ({columns_sql}) VALUES ({placeholders})"
            # cursor.execute(sql, values)
            columns = list(row_dict.keys())
            values = [row_dict[col] for col in columns]

            placeholders = ', '.join(['%s'] * len(columns))
            columns_sql = ', '.join(columns)

            update_clause = ', '.join([f"{col}=EXCLUDED.{col}" for col in columns if col != 'id'])

            sql = f"""
            INSERT INTO transacts ({columns_sql})
            VALUES ({placeholders})
            ON CONFLICT (tscode) DO UPDATE
            SET {update_clause}
            """
            cursor.execute(sql, values)

    else:
        return jsonify({"message": f"{filename} was detected to not be a csv or xlsx, no importing could be done"}), 400 
    
    # commit changes
    conn.commit()

    return jsonify({"message": f"{filename} uploaded successfully"}), 200

if __name__ == "__main__":
    app.run(port=5000)