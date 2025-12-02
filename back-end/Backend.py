# Need to install:
# pip install psycopg2-binary scikit-learn pandas numpy openpyxl flask-cors python-dotenv

from dotenv import load_dotenv
from flask import Flask, request, jsonify
from flask_cors import CORS
import os
from datetime import datetime, date
import bcrypt
import io
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import io
import csv
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import RealDictCursor

# ML
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import OneHotEncoder
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.metrics import roc_auc_score
from catboost import CatBoostClassifier, Pool


# -------------------------------------------------
# Setup / DB connection
# -------------------------------------------------
load_dotenv()
app = Flask(__name__)

ALLOWED_ORIGINS = ["http://localhost:5173", "http://127.0.0.1:5173"]
CORS(
    app,
    resources={r"/*": {"origins": ALLOWED_ORIGINS}},
    supports_credentials=True,
    methods=["GET", "POST", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
)


@app.after_request
def add_cors_headers(resp):
    origin = request.headers.get("Origin")
    if origin in ALLOWED_ORIGINS:
        resp.headers["Access-Control-Allow-Origin"] = origin
        resp.headers["Vary"] = "Origin"
        resp.headers["Access-Control-Allow-Credentials"] = "true"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
        resp.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    return resp


@app.route("/<path:any_path>", methods=["OPTIONS"])
def preflight(any_path):
    return ("", 204)


conn = psycopg2.connect(
    dbname=os.getenv("DB_NAME"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASSWORD"),
    host=os.getenv("DB_HOST"),
    port=os.getenv("DB_PORT"),
)
conn.autocommit = True
cursor = conn.cursor()

# -------------------------------------------------
# Schema (matches your manual load)
# -------------------------------------------------
cursor.execute("""
CREATE TABLE IF NOT EXISTS transacts (
    pscode TEXT,
    tscode TEXT PRIMARY KEY,
    uscode TEXT,
    screenresult TEXT,
    screenvendor TEXT,
    dnumnsf INTEGER,
    dnumlate INTEGER,
    davgdayslate INTEGER,
    sevicted TEXT,
    smoveoutreason TEXT,
    drentwrittenoff NUMERIC,
    dnonrentwrittenoff NUMERIC,
    damoutcollections NUMERIC,
    srenewed TEXT,
    srent NUMERIC,
    sfulfilledterm TEXT,
    dincome NUMERIC,
    dtleasefrom DATE,
    dtleaseto DATE,
    dtmovein DATE,
    dtmoveout DATE,
    dwocount INTEGER,
    dtroomearlyout DATE,
    sempcompany TEXT,
    sempposition TEXT,
    sflex TEXT,
    sleap TEXT,
    sprevzip TEXT,
    daypaid INTEGER,
    spaymentsource TEXT,
    dpaysourcechange INTEGER
);
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS screening (
    appcredid TEXT,
    appcreddate DATE,
    appid TEXT,
    category TEXT,
    city TEXT,
    companycode TEXT,
    companyname TEXT,
    creditrun BOOLEAN,
    date DATE,
    propertyid TEXT,
    policy TEXT,
    posemployment TEXT,
    poshousing TEXT,
    propname TEXT,
    reasonone TEXT,
    reasontwo TEXT,
    reasonthree TEXT,
    rentownhist TEXT,
    origscore TEXT,
    finscore TEXT,
    scorecat TEXT,
    scoremodel INTEGER,
    marketsource TEXT,
    state TEXT,
    zip TEXT,
    age NUMERIC,
    currempmon NUMERIC,
    currempyear NUMERIC,
    currresmon NUMERIC,
    currresyear NUMERIC,
    income NUMERIC,
    primincome NUMERIC,
    addincome NUMERIC,
    riskscore NUMERIC,
    prevempmon NUMERIC,
    prevempyear NUMERIC,
    prevresmon NUMERIC,
    prevresyear NUMERIC,
    rent NUMERIC,
    rentincratio NUMERIC,
    debtincratio NUMERIC,
    debtcredratio NUMERIC,
    voyappcode TEXT PRIMARY KEY,
    voypropname TEXT,
    voypropcode TEXT,
    hascpmess TEXT,
    checkmes1 TEXT,
    checkmes2 TEXT,
    hasconsstmt TEXT,
    studdebt TEXT,
    meddebt TEXT,
    totscordebt NUMERIC,
    totdebt NUMERIC,
    itemrev1 TEXT,
    itemrev2 TEXT,
    itemrev3 TEXT,
    revrepack TEXT,
    appid2 TEXT,
    appscore TEXT,
    appmoninc TEXT,
    apptotdebt NUMERIC,
    avgriskscore NUMERIC,
    twnreport TEXT,
    appstatus TEXT
);
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS meta_updates (
    id SERIAL PRIMARY KEY,
    updated_at TIMESTAMP NOT NULL DEFAULT NOW()
);
""")
conn.commit()

# Helpful indexes for WHERE clauses in KPI queries / filters
def ensure_indexes():
    # NOTE: CREATE INDEX IF NOT EXISTS is cheap if it already exists.
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_transacts_pscode ON transacts (pscode);"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_transacts_screenresult ON transacts (screenresult);"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_transacts_sevicted ON transacts (sevicted);"
    )
    # dates used to bucket: these help planner even with coalesce/date_trunc
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_transacts_dates_movein ON transacts (dtmovein);"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_transacts_dates_leasefrom ON transacts (dtleasefrom);"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_transacts_dates_roomout ON transacts (dtroomearlyout);"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_transacts_dates_leaseto ON transacts (dtleaseto);"
    )
    conn.commit()

ensure_indexes()

# -------------------------------------------------
# Users / Auth
# -------------------------------------------------
def ensure_users_table():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            email VARCHAR(255) UNIQUE NOT NULL,
            password_hash VARCHAR(255) NOT NULL,
            role VARCHAR(50) DEFAULT 'end-user',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conn.commit()

ensure_users_table()

@app.route('/register', methods=['POST', 'OPTIONS'])
def register():
    if request.method == 'OPTIONS':
        return jsonify({'message': 'OK'}), 200

    data = request.get_json()
    name = data.get('name')
    email = data.get('email')
    password = data.get('password')

    if not name or not email or not password:
        return jsonify({'error': 'Information missing'}), 400

    hashed_pw = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf8')

    try:
        cursor.execute("""
            INSERT INTO users (name, email, password_hash)
            VALUES (%s, %s, %s)
            RETURNING id, name, email, role
        """, (name, email, hashed_pw))
        conn.commit()
        new_user = cursor.fetchone()
        return jsonify({
            'id': new_user[0],
            'name': new_user[1],
            'email': new_user[2],
            'role': new_user[3]
        }), 201
    except psycopg2.Error as e:
        conn.rollback()
        if 'unique' in str(e).lower():
            return jsonify({'error': 'Email exists'}), 409
        return jsonify({'error': 'Database error'}), 500


@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')

    if not email or not password:
        return jsonify({'error': 'Missing either email or password'}), 400

    cursor.execute(
        "SELECT id, name, email, password_hash, role FROM users WHERE email = %s",
        (email,)
    )
    user = cursor.fetchone()

    if not user:
        return jsonify({'error': 'Invalid credentials'}), 401

    if bcrypt.checkpw(password.encode('utf-8'), user[3].encode('utf-8')):
        return jsonify({
            'id': user[0],
            'name': user[1],
            'email': user[2],
            'role': user[4]
        }), 200

    return jsonify({'error': 'Invalid credentials'}), 401


# -------------------------------------------------
# Helpers
# -------------------------------------------------
PRIMARY_KEY = None
DATE_COLUMNS = ['dtleasefrom', 'dtleaseto', 'dtmovein', 'dtmoveout', 'dtroomearlyout']


def _normalize_row(row: dict):
    """Trim keys/values, empty->None, and parse dates (mm/dd/yyyy)."""
    row = {str(k).strip().lower(): (str(v).strip() if (v is not None and str(v).strip() != '') else None)
           for k, v in row.items()}
    for col in DATE_COLUMNS:
        if row.get(col):
            try:
                row[col] = datetime.strptime(row[col], "%m/%d/%Y").date()
            except Exception:
                try:
                    # Try ISO style if already clean
                    if isinstance(row[col], (datetime, date)):
                        row[col] = row[col]
                    else:
                        row[col] = datetime.fromisoformat(str(row[col])).date()
                except Exception:
                    row[col] = None
    return row

# --- Mapping of names for screening data ---
SCREEN_MAPPING = {
    "screening": {
        "Applicant Credit Applicant ID": "appcredid",
        "Applicant Credit Date": "appcreddate",
        "Applicant ID": "appid",
        "Category": "category",
        "City": "city",
        "Company Code": "companycode",
        "Company Name": "companyname",
        "Credit Run": "creditrun",
        "Date": "date",
        "Property ID": "propertyid",
        "Policy": "policy",
        "Positive Employment": "posemployment",
        "Positive Housing": "poshousing",
        "Property Name": "propname",
        "Reason 1": "reasonone",
        "Reason 2": "reasontwo",
        "Reason 3": "reasonthree",
        "Rent Own History": "rentownhist",
        "Original Score": "origscore",
        "Final Score": "finscore",
        "Score Category": "scorecat",
        "Score Model": "scoremodel",
        "Market Source": "marketsource",
        "State": "state",
        "Zip": "zip",
        "Age" : "age",
        "Current Emp (Months)": "currempmon",
        "Current Emp (Years)": "currempyear",
        "Current Res (Months)": "currresmon",
        "Current Res (Years)": "currresyear",
        "Income": "income",
        "Primary Income": "primincome",
        "Additional Income": "addincome",
        "Risk Score": "riskscore",
        "Previous Emp (Months)": "prevempmon",
        "Previous Emp (Years)": "prevempyear",
        "Previous Res (Months)": "prevresmon",
        "Previous Res (Years)": "prevresyear",
        "Rent": "rent",
        "Rent To Income Ratio (%)": "rentincratio",
        "Debt To Income Ratio (%)": "debtincratio",
        "Debt To Credit Ratio (%)": "debtcredratio",
        "Voyager Applicant Code": "voyappcode",
        "Voyager Property Name": "voypropname",
        "Voyager Property Code" : "voypropcode",
        "Has CheckPoint Msgs": "hascpmess",
        "Checkpoint Message 1": "checkmes1",
        "Checkpoint Message 2": "checkmes2",
        "Has Consumer Stmt": "hasconsstmt",
        "Student Debt": "studdebt",
        "Medical Debt": "meddebt",
        "Total Scorable Debt": "totscordebt",
        "Total Debt": "totdebt",
        "Item To Review 1": "itemrev1",
        "Item To Review 2": "itemrev2",
        "Item To Review 3": "itemrev3",
        "Review Report Acknowledgement": "revrepack",
        "Application ID": "appid2",
        "Application Score": "appscore",
        "Application Monthly Income": "appmoninc",
        "Application Total Debt (Policy)": "apptotdebt",
        "Avg Risk Score": "avgriskscore",
        "TWN Report Found": "twnreport",
        "Applicant Status": "appstatus"
    }
}

# Lowercase all keys in the mapping for screening
SCREEN_MAPPING["screening"] = {
    k.lower(): v for k, v in SCREEN_MAPPING["screening"].items()
}

# -------------------------------------------------
# Upload route for data
# -------------------------------------------------
@app.route("/upload", methods=["POST"])
def upload_file():
    file = None
    dataName = None
    if "transact" in request.files:
        file = request.files["transact"]
        dataName = "transacts"
        PRIMARY_KEY = "tscode"
    elif "screening" in request.files:
        file = request.files["screening"]
        dataName = "screening"
        PRIMARY_KEY = "voyappcode"
        col_map = SCREEN_MAPPING.get(dataName, {})

        def map_columns(row):
            mapped = {}
            for key, value in row.items():
                mapped[col_map.get(key.strip(), key)] = value
            return mapped

    else:
        return jsonify({"error": "No file uploaded"}), 400
    
    filename = file.filename
    content = file.read()  # bytes

    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        try:
            csv_file = io.StringIO(content.decode("utf-8-sig"))
        except UnicodeDecodeError:
            return jsonify({"message": f"{filename} is not UTF-8 encoded"}), 400

        # Skip first 5 rows (headers/noise)
        # for _ in range(5):
        #     next(csv_file, None)

        reader = csv.DictReader(csv_file)
        reader.fieldnames = [h.strip().lower() for h in reader.fieldnames]

        # Prepare upsert dynamically based on first row
        first = next(reader, None)
        if first is None:
            return jsonify({"message": "No rows detected after header"}), 400

        first = _normalize_row(first)
        if dataName == "screening":
            first = map_columns(first)   # apply SCREEN_MAPPING here
        columns = list(first.keys())
        placeholders = ', '.join(['%s'] * len(columns))
        colsql = ', '.join(columns)
        update_clause = ', '.join([f"{c}=EXCLUDED.{c}" for c in columns if c != PRIMARY_KEY])

        sql = f"INSERT INTO {dataName} ({colsql}) VALUES ({placeholders}) ON CONFLICT ({PRIMARY_KEY}) DO UPDATE SET {update_clause}"

        batch, batch_size = [], 1000

        def add_row(r):
            r = _normalize_row(r)
            if dataName == "screening":
                r = map_columns(r)
            # print("Primary Key:", PRIMARY_KEY)
            # print("Row keys:", list(r.keys()))
            if not r.get(PRIMARY_KEY):
                return
            batch.append([r.get(c) for c in columns])

        add_row(first)
        for r in reader:
            add_row(r)
            if len(batch) >= batch_size:
                cursor.executemany(sql, batch)
                batch.clear()
        if batch:
            print("Columns:", columns)
            print("First batch row:", batch[0] if batch else None)
            print("SQL:", sql)
            cursor.executemany(sql, batch)

    elif ext == ".xlsx":
        xlsx = io.BytesIO(content)
        wb = load_workbook(xlsx, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        rows = rows[5:]  # skip first 5 rows
        if not rows:
            return jsonify({"message": "No data rows detected"}), 400
        headers = [str(h).strip().lower() for h in rows[0]]
        for row in rows[1:]:
            rd = {k: v for k, v in zip(headers, row)}
            rd = _normalize_row(rd)
            if dataName == "screening":
                rd = map_columns(rd)
            if not rd.get(PRIMARY_KEY):
                continue
            cols = list(rd.keys())
            placeholders = ', '.join(['%s'] * len(cols))
            colsql = ', '.join(cols)
            update_clause = ', '.join([f"{c}=EXCLUDED.{c}" for c in cols if c != PRIMARY_KEY])
            sql = f"INSERT INTO transacts ({colsql}) VALUES ({placeholders}) ON CONFLICT ({PRIMARY_KEY}) DO UPDATE SET {update_clause}"
            cursor.execute(sql, [rd.get(c) for c in cols])
    else:
        return jsonify({"message": "Unsupported file type"}), 400

    # Touch meta_updates
    cursor.execute("INSERT INTO meta_updates (updated_at) VALUES (NOW())")
    return jsonify({"message": f"{filename} uploaded successfully"}), 200

def _bucketsql():
    # Which month a row counts toward (for grouping)
    return """date_trunc('month',
              coalesce(dtmovein, dtleasefrom, dtroomearlyout, dtleaseto)
            )::date"""


def _clean_pscode_sql():
    # Strip trailing ".0" from pscode
    return "regexp_replace(pscode, '\\.0+$', '')"


def _build_filter_sql(params, allow_dates=True):
    where = []
    vals = []

    # date window: frontend passes YYYY-MM
    if allow_dates:
        start = params.get("start")
        end = params.get("end")
        if start:
            where.append(f"{_bucketsql()} >= date_trunc('month', %s::date)")
            vals.append(f"{start}-01")
        if end:
            where.append(f"{_bucketsql()} <= (date_trunc('month', %s::date) + interval '1 month - 1 day')::date")
            vals.append(f"{end}-01")

    # multi-pscode
    pscodes = params.getlist("pscode") if hasattr(params, "getlist") else params.get("pscode")
    if pscodes:
        if isinstance(pscodes, str):
            pscodes = [pscodes]
        where.append(f"{_clean_pscode_sql()} = ANY(%s)")
        vals.append(pscodes)

    # screen result
    screen = params.get("screenresult")
    if screen:
        where.append("screenresult = %s")
        vals.append(screen)

    # collections filter
    collections = params.get("collections")
    if collections == "with":
        where.append("coalesce(damoutcollections,0) > 0")
    elif collections == "without":
        where.append("coalesce(damoutcollections,0) = 0")

    # eviction filter
    ev = params.get("evicted")
    if ev in ("Yes", "No"):
        where.append("sevicted = %s")
        vals.append(ev)

    return ("WHERE " + " AND ".join(where)) if where else "", vals


# -------------------------------------------------
# /filters/options
# -------------------------------------------------
@app.route("/filters/options")
def filter_options():
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"""
                SELECT DISTINCT {_clean_pscode_sql()} AS pcode_clean
                FROM transacts
                WHERE pscode IS NOT NULL
                ORDER BY pcode_clean
            """)
            props = [r["pcode_clean"] for r in cur.fetchall()]

            cur.execute("""
                SELECT DISTINCT screenresult
                FROM transacts
                WHERE screenresult IS NOT NULL
                ORDER BY screenresult
            """)
            screens = [r["screenresult"] for r in cur.fetchall()]

        return jsonify({"pscodes": props, "screenresults": screens})
    except Exception as e:
        print(f"Error in filter_options: {str(e)}")
        return jsonify({"pscodes": [], "screenresults": []}), 200


# -------------------------------------------------
# KPI queries with positive dollar magnitudes
# -------------------------------------------------
def _query_snapshot(where_clause, vals):
    """
    Aggregate portfolio-level KPIs over the filtered window.

    collections_exposure  = sum of dollars sent to collections
    dollars_delinquent    = total delinquent exposure
                            = collections_exposure + rent/non-rent write-offs
    """
    q = f"""
        WITH base AS (
            SELECT {_bucketsql()} AS month_key,
                   dnumlate,
                   dnumnsf,
                   damoutcollections,
                   drentwrittenoff,
                   dnonrentwrittenoff
            FROM transacts
            {where_clause}
        )
        SELECT
          COUNT(*) AS total_rows,
          COALESCE(SUM(CASE WHEN dnumlate > 0 THEN 1 ELSE 0 END),0)::float
            / NULLIF(COUNT(*),0) AS pct_late_payers,
          COALESCE(SUM(dnumnsf),0) AS nsf_count,
          -- dollars currently in collections
          ABS(COALESCE(SUM(damoutcollections), 0)) AS collections_exposure,
          -- total delinquent exposure = collections + write-offs
          ABS(
            COALESCE(SUM(damoutcollections), 0)
            + COALESCE(SUM(drentwrittenoff), 0)
            + COALESCE(SUM(dnonrentwrittenoff), 0)
          ) AS dollars_delinquent
        FROM base;
    """
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(q, vals)
        row = cur.fetchone()
    return row



def _query_timeseries(where_clause, vals):
    """
    Monthly time-series for portfolio KPIs.

    collections_exposure  = sum of dollars sent to collections
    dollars_delinquent    = total delinquent exposure
                            = collections_exposure + rent/non-rent write-offs
    """
    q = f"""
        WITH base AS (
            SELECT {_bucketsql()} AS month_key,
                   dnumlate,
                   dnumnsf,
                   damoutcollections,
                   drentwrittenoff,
                   dnonrentwrittenoff
            FROM transacts
            {where_clause}
        )
        SELECT
          month_key,
          COALESCE(SUM(CASE WHEN dnumlate > 0 THEN 1 ELSE 0 END),0)::float
            / NULLIF(COUNT(*),0) AS pct_late_payers,
          COALESCE(SUM(dnumnsf),0) AS nsf_count,
          -- dollars currently in collections
          ABS(COALESCE(SUM(damoutcollections), 0)) AS collections_exposure,
          -- total delinquent exposure = collections + write-offs
          ABS(
            COALESCE(SUM(damoutcollections), 0)
            + COALESCE(SUM(drentwrittenoff), 0)
            + COALESCE(SUM(dnonrentwrittenoff), 0)
          ) AS dollars_delinquent
        FROM base
        GROUP BY month_key
        ORDER BY month_key;
    """
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(q, vals)
        return cur.fetchall()



# -------------------------------------------------
# /kpis/snapshot
# -------------------------------------------------
@app.route("/kpis/snapshot")
def kpi_snapshot():
    try:
        where1, vals1 = _build_filter_sql(request.args, allow_dates=True)
        row = _query_snapshot(where1, vals1)

        if not row or (row["total_rows"] or 0) == 0:
            where2, vals2 = _build_filter_sql(request.args, allow_dates=False)
            row = _query_snapshot(where2, vals2)

        if not row or (row["total_rows"] or 0) == 0:
            return jsonify({
                "pct_late_payers": 0.0,
                "nsf_count": 0,
                "collections_exposure": 0.0,
                "dollars_delinquent": 0.0
            })

        return jsonify({
            "pct_late_payers": float(row["pct_late_payers"] or 0.0),
            "nsf_count": int(row["nsf_count"] or 0),
            "collections_exposure": float(row["collections_exposure"] or 0.0),
            "dollars_delinquent": float(row["dollars_delinquent"] or 0.0)
        })
    except Exception as e:
        print(f"Error in kpi_snapshot: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "pct_late_payers": 0.0,
            "nsf_count": 0,
            "collections_exposure": 0.0,
            "dollars_delinquent": 0.0
        }), 200


# -------------------------------------------------
# /kpis/timeseries
# -------------------------------------------------
@app.route("/kpis/timeseries")
def kpi_timeseries():
    try:
        where1, vals1 = _build_filter_sql(request.args, allow_dates=True)
        rows = _query_timeseries(where1, vals1)

        if len(rows) == 0:
            where2, vals2 = _build_filter_sql(request.args, allow_dates=False)
            rows = _query_timeseries(where2, vals2)

        out = []
        for r in rows:
            mk = r["month_key"]
            out.append({
                "month": mk.strftime("%Y/%m") if mk else None,
                "pct_late_payers": float(r["pct_late_payers"] or 0.0),
                "nsf_count": int(r["nsf_count"] or 0),
                "collections_exposure": float(r["collections_exposure"] or 0.0),
                "dollars_delinquent": float(r["dollars_delinquent"] or 0.0),
            })
        return jsonify(out)
    except Exception as e:
        print(f"Error in kpi_timeseries: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify([]), 200


# -------------------------------------------------
# Feature importance with proper missing value handling
# + in-memory caching so we don't retrain every request
# -------------------------------------------------
_PREDICTOR_COLS = [
    "dnumnsf", "dnumlate", "davgdayslate", "srent", "dincome",
    "daypaid", "dpaysourcechange",
    "screenresult", "screenvendor", "sflex", "sleap", "spaymentsource", "sprevzip"
]

_HUMAN_NAME = {
    "dnumnsf": "NSF count",
    "dnumlate": "Late payment count",
    "davgdayslate": "Average days late",
    "srent": "Monthly rent",
    "dincome": "Reported income",
    "daypaid": "Day-of-month paid",
    "dpaysourcechange": "Payment source changed",
    "screenresult": "Screen result",
    "screenvendor": "Screen vendor",
    "sflex": "Flex program",
    "sleap": "LEAP program",
    "spaymentsource": "Payment source",
    "sprevzip": "Previous ZIP code",
}

def _humanize_feature_name(raw: str) -> str:
    if "_" in raw:
        prefix, val = raw.split("_", 1)
        if prefix in _HUMAN_NAME:
            return f"{_HUMAN_NAME[prefix]}: {val.replace('_', ' ')}"
    return _HUMAN_NAME.get(raw, raw)

def _train_rf_with_imputation(X: pd.DataFrame, y: pd.Series):
    """
    Train Random Forest with proper missing value handling via imputation.
    Returns (pipeline, auc_score, feature_names) or (None, None, None) on failure.
    """
    try:
        num_cols = X.select_dtypes(include=[np.number]).columns.tolist()
        cat_cols = [c for c in X.columns if c not in num_cols]

        transformers = []

        if num_cols:
            num_pipeline = Pipeline([
                ('imputer', SimpleImputer(strategy='median')),
            ])
            transformers.append(('num', num_pipeline, num_cols))

        if cat_cols:
            cat_pipeline = Pipeline([
                ('imputer', SimpleImputer(strategy='constant', fill_value='missing')),
                ('encoder', OneHotEncoder(handle_unknown='ignore', sparse_output=False))
            ])
            transformers.append(('cat', cat_pipeline, cat_cols))

        if not transformers:
            return None, None, None

        preprocessor = ColumnTransformer(
            transformers=transformers,
            remainder='drop'
        )

        # Train/test split
        try:
            X_train, X_test, y_train, y_test = train_test_split(
                X, y, test_size=0.25, random_state=42, stratify=y
            )
        except ValueError:
            X_train, X_test, y_train, y_test = train_test_split(
                X, y, test_size=0.25, random_state=42
            )

        rf = RandomForestClassifier(
            n_estimators=200,
            max_depth=10,
            min_samples_leaf=5,
            max_features='sqrt',
            n_jobs=-1,
            random_state=42,
            class_weight='balanced'
        )

        pipeline = Pipeline([
            ('preprocessor', preprocessor),
            ('classifier', rf)
        ])

        pipeline.fit(X_train, y_train)

        y_pred_proba = pipeline.predict_proba(X_test)[:, 1]
        auc = roc_auc_score(y_test, y_pred_proba)

        feature_names = []
        if num_cols:
            feature_names.extend(num_cols)
        if cat_cols:
            cat_feature_names = (
                pipeline.named_steps['preprocessor']
                .named_transformers_['cat']
                .named_steps['encoder']
                .get_feature_names_out(cat_cols)
            )
            feature_names.extend(cat_feature_names)

        return pipeline, auc, feature_names

    except Exception as e:
        print(f"RF training error: {str(e)}")
        return None, None, None

# simple in-memory caches
_FEATURE_IMPORTANCE_CACHE = {
    "last_meta_ts": None,
    "payload": None,
}

# NEW: cache for transaction-model eviction scores
_TRANSACTION_MODEL_CACHE = {
    "last_meta_ts": None,
    "payload": None,
}

_SCREENING_MODEL_CACHE = {
    "last_meta_ts": None,
    "payload": None,
}


def _map_flag(x):
    """
    Normalize a yes/no-style flag into {1, 0, NaN}.
    Used to build the sevicted label for the transaction model.
    """
    import math

    if x is None:
        return float("nan")
    try:
        if isinstance(x, float) and math.isnan(x):
            return float("nan")
    except TypeError:
        pass

    val = str(x).strip().upper()
    if val in ("Y", "YES", "1", "TRUE", "T"):
        return 1
    if val in ("N", "NO", "0", "FALSE", "F"):
        return 0
    return float("nan")


def _latest_meta_ts():
    """Return latest updated_at from meta_updates, or None if table empty."""
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT MAX(updated_at) FROM meta_updates;")
            row = cur.fetchone()
            return row[0] if row else None
    except Exception:
        return None

def _compute_feature_importance_payload():
    """
    Run the expensive pandas + RF pipeline once and return the JSON payload.
    """
    q = """
        SELECT
          sevicted,
          dnumnsf, dnumlate, davgdayslate, srent, dincome,
          daypaid, dpaysourcechange,
          screenresult, screenvendor, sflex, sleap, spaymentsource, sprevzip
        FROM transacts
        WHERE sevicted IS NOT NULL
    """
    df = pd.read_sql(q, conn)

    # sanity checks
    if df.empty or len(df) < 50:
        return {
            "auc": None,
            "top_features": []
        }

    evicted_counts = df["sevicted"].astype(str).str.upper().value_counts()
    if len(evicted_counts) < 2:
        return {
            "auc": None,
            "top_features": []
        }

    y = (df["sevicted"].astype(str).str.upper() == "YES").astype(int)

    X = df.drop(columns=["sevicted"])
    X = X[[c for c in _PREDICTOR_COLS if c in X.columns]]

    # remove columns that are entirely missing
    X = X.dropna(axis=1, how='all')
    if X.empty or len(X.columns) == 0:
        return {
            "auc": None,
            "top_features": []
        }

    pipeline, auc, feature_names = _train_rf_with_imputation(X, y)
    if pipeline is None or feature_names is None:
        return {
            "auc": None,
            "top_features": []
        }

    feature_importances = pipeline.named_steps['classifier'].feature_importances_
    feature_importance_pairs = list(zip(feature_names, feature_importances))
    feature_importance_pairs.sort(key=lambda x: x[1], reverse=True)

    top_features = [
        {
            "feature": _humanize_feature_name(name),
            "importance": float(importance)
        }
        for name, importance in feature_importance_pairs[:20]
    ]

    return {
        "auc": float(auc),
        "top_features": top_features
    }

def _compute_transaction_model_payload():
    """
    Train a CatBoost model on historical transaction data to predict eviction,
    then return 0–100 eviction risk scores for the 2024+ cohort,
    grouped by property code, with per-tenant top driver features.

    Now includes:
      - daypaid (numeric)
      - dpaysourcechange (numeric)
      - spaymentsource (categorical)
    """
    q = """
    SELECT
        pscode,
        tscode,
        uscode,
        dtleasefrom,
        dtleaseto,
        dtmovein,
        dtmoveout,
        dnumnsf,
        dnumlate,
        davgdayslate,
        drentwrittenoff,
        dnonrentwrittenoff,
        damoutcollections,
        srenewed,
        srent,
        sfulfilledterm,
        dincome,
        daypaid,
        spaymentsource,
        dpaysourcechange,
        sevicted
    FROM transacts
    WHERE pscode IS NOT NULL
      AND tscode IS NOT NULL
      AND dtmovein IS NOT NULL;
    """

    df = pd.read_sql(q, conn)

    if df.empty or "sevicted" not in df.columns:
        return {}

    # Normalize dates
    for col in ["dtleasefrom", "dtleaseto", "dtmovein", "dtmoveout"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    # Binary label from sevicted
    def _map_flag_local(x):
        if x is None:
            return np.nan
        s = str(x).strip().lower()
        if s in ("yes", "y", "1", "true"):
            return 1
        if s in ("no", "n", "0", "false"):
            return 0
        return np.nan

    df["sevicted_flag"] = df["sevicted"].apply(_map_flag_local)
    df = df[~df["sevicted_flag"].isna()].copy()
    if len(df) < 100:
        # Not enough labeled data to train a reasonable model
        return {}

    df["label"] = df["sevicted_flag"].astype(int)

    # Lease start: prefer dtleasefrom, fallback to dtmovein
    df["lease_start"] = df["dtleasefrom"].where(
        df["dtleasefrom"].notna(),
        df["dtmovein"]
    )
    df = df[~df["lease_start"].isna()].copy()
    df["start_year"] = df["lease_start"].dt.year

    # Train on <= 2023, score on >= 2024
    train_df = df[df["start_year"] <= 2023].copy()
    test_df = df[df["start_year"] >= 2024].copy()

    # Further restrict scored cohort to move-ins in 2024+
    min_movein = pd.Timestamp("2024-01-01")
    if "dtmovein" in test_df.columns:
        test_df = test_df[test_df["dtmovein"] >= min_movein].copy()

    # If we don't have both sides, bail out gracefully
    if train_df.empty or test_df.empty:
        return {}

    # ------------------------------------------------------------------
    # Feature engineering shared by train/test
    # ------------------------------------------------------------------
    for df_ in (train_df, test_df):
        # Payment behavior ratios
        df_["late_ratio"] = (df_["dnumlate"] / df_["dnumnsf"]).replace(
            [np.inf, -np.inf], np.nan
        )
        df_["wo_total"] = (df_["drentwrittenoff"] + df_["dnonrentwrittenoff"]).fillna(0)
        df_["collections_flag"] = (df_["damoutcollections"].fillna(0) > 0).astype(int)
        df_["renewed_flag"] = df_["srenewed"].astype(str).str.lower().isin(
            ["yes", "y", "1", "true"]
        )
        df_["fulfilled_flag"] = df_["sfulfilledterm"].astype(str).str.lower().isin(
            ["yes", "y", "1", "true"]
        )

        # Income / rent ratios
        df_["rent_to_income"] = np.where(
            df_["dincome"].notna() & (df_["dincome"] > 0),
            df_["srent"] / df_["dincome"],
            np.nan,
        )

        # Tenure as numeric
        df_["tenure_days"] = (
            df_["dtmoveout"].fillna(pd.Timestamp("today")) - df_["dtmovein"]
        ).dt.days

    # ------------------------------------------------------------------
    # Features the transaction model is allowed to look at
    # (no write-off / collections columns to avoid label leakage)
    # ------------------------------------------------------------------
    candidate_cols = [
        "dnumnsf",
        "dnumlate",
        "davgdayslate",
        "srent",
        "dincome",
        "late_ratio",
        "rent_to_income",
        "tenure_days",
        "daypaid",
        "dpaysourcechange",
        "spaymentsource",
    ]

    # Only keep columns actually present
    candidate_cols = [c for c in candidate_cols if c in train_df.columns]

    # Do not leak the label or sevicted text itself
    leakage_cols = ["sevicted", "sevicted_flag", "label"]
    FEATURES = [c for c in candidate_cols if c not in leakage_cols]

    if not FEATURES:
        return {}

    # ------------------------------------------------------------------
    # Driver candidates – subset of FEATURES, with "worse than baseline"
    # direction encoded per feature.
    # ------------------------------------------------------------------
    driver_specs = {
        "dnumlate": {
            "label": "Late payment count",
            "direction": "high",  # more late payments = worse
        },
        "dnumnsf": {
            "label": "NSF count",
            "direction": "high",  # more NSFs = worse
        },
        "davgdayslate": {
            "label": "Average days late",
            "direction": "high",  # more days late = worse
        },
        "rent_to_income": {
            "label": "Rent-to-income (%)",
            "direction": "high",  # higher rent burden = worse
        },
        "tenure_days": {
            "label": "Tenure length (days)",
            "direction": "low",  # shorter tenure = worse
        },
        "daypaid": {
            "label": "Day-of-month paid",
            "direction": "high",  # later in month = worse
        },
        "dpaysourcechange": {
            "label": "Payment-source changes",
            "direction": "high",  # more changes = worse
        },
        "spaymentsource": {
            "label": "Payment source",
            "direction": "category",  # different than low-risk norm = worse
        },
    }

    # Safety: only keep drivers that are actually in the model feature set
    driver_specs = {k: v for k, v in driver_specs.items() if k in FEATURES}

    low_risk_train = train_df[train_df["label"] == 0]
    baseline = {}
    spread = {}

    if not low_risk_train.empty:
        for col, spec in driver_specs.items():
            if col not in low_risk_train.columns:
                continue

            # Categorical baseline (e.g., spaymentsource)
            if isinstance(spec, dict) and spec.get("direction") == "category":
                series_raw = low_risk_train[col].dropna()
                if series_raw.empty:
                    continue
                mode_val = series_raw.mode()
                if not mode_val.empty:
                    baseline[col] = str(mode_val.iloc[0])
                    spread[col] = None
                continue

            # Numeric baseline (median + std)
            series = pd.to_numeric(low_risk_train[col], errors="coerce").dropna()
            if series.empty:
                continue
            baseline[col] = float(series.median())
            std_val = float(series.std(ddof=0))
            spread[col] = std_val if std_val > 0 else None

    # ------------------------------------------------------------------
    # Model training
    # ------------------------------------------------------------------
    X_train = train_df[FEATURES].copy()
    y_train = train_df["label"].copy()
    X_test = test_df[FEATURES].copy()
    y_test = test_df["label"].copy()  # target for eval set

    # Simple numeric / categorical imputation
    for col in FEATURES:
        if X_train[col].dtype.kind in "biufc":  # numeric
            median = X_train[col].median()
            X_train[col] = X_train[col].fillna(median)
            X_test[col] = X_test[col].fillna(median)
        else:
            mode = X_train[col].mode(dropna=True)
            fill_value = mode.iloc[0] if not mode.empty else ""
            X_train[col] = X_train[col].fillna(fill_value)
            X_test[col] = X_test[col].fillna(fill_value)

    # Mark categorical features for CatBoost
    cat_features_idx = []
    for idx, col in enumerate(FEATURES):
        if X_train[col].dtype.kind not in "biufc":
            cat_features_idx.append(idx)
            X_train[col] = X_train[col].astype("string")
            X_test[col] = X_test[col].astype("string")

    train_pool = Pool(X_train, label=y_train, cat_features=cat_features_idx)
    test_pool = Pool(X_test, label=y_test, cat_features=cat_features_idx)

    model = CatBoostClassifier(
        loss_function="Logloss",
        eval_metric="AUC",
        depth=6,
        learning_rate=0.05,
        iterations=800,
        l2_leaf_reg=5,
        random_seed=42,
        auto_class_weights="Balanced",
        od_type="Iter",
        od_wait=200,
        verbose=False,
    )

    model.fit(train_pool, eval_set=test_pool, use_best_model=True)

    # ------------------------------------------------------------------
    # Global top drivers for the transaction model (feature importance)
    # ------------------------------------------------------------------
    try:
        importances = model.get_feature_importance(train_pool)

        # Exclude features you don't want to show in the global top-drivers strip
        EXCLUDED_GLOBAL_TX_DRIVERS = {"fulfilled_flag", "renewed_flag"}

        pairs = [
            (name, imp)
            for name, imp in zip(FEATURES, importances)
            if name not in EXCLUDED_GLOBAL_TX_DRIVERS
        ]
        pairs.sort(key=lambda x: x[1], reverse=True)

        global_tx_drivers = []
        for feat_name, imp in pairs[:3]:
            global_tx_drivers.append(
                {
                    "feature_key": feat_name,
                    "feature_label": _pretty_transaction_feature_label(feat_name),
                    "importance": float(imp),
                }
            )

        _TRANSACTION_MODEL_CACHE["global_drivers"] = global_tx_drivers
    except Exception as e:
        print(f"Could not compute global transaction drivers: {e}")
        _TRANSACTION_MODEL_CACHE["global_drivers"] = []

    y_proba_test = model.predict_proba(test_pool)[:, 1]
    risk_score_0_100 = (y_proba_test * 100).round(1)

    test_df = test_df.copy()
    test_df["eviction_risk_score"] = risk_score_0_100

    # Build property -> tenants mapping (only 2024+ cohort)
    out = {}
    for idx, row in test_df.iterrows():
        pscode = row.get("pscode")
        if not pscode:
            continue

        if pscode not in out:
            out[pscode] = []

        lease_start_val = row.get("lease_start")
        dtmovein_val = row.get("dtmovein")
        dtmoveout_val = row.get("dtmoveout")

        # Per-tenant top drivers for the transaction model
        top_drivers = _compute_top_drivers(
            row,
            driver_specs=driver_specs,
            baseline=baseline,
            spread=spread,
            max_drivers=3,
        )

        out[pscode].append(
            {
                "tscode": row.get("tscode"),
                "uscode": row.get("uscode"),
                "lease_start": lease_start_val.date().isoformat()
                if pd.notna(lease_start_val)
                else None,
                "dtmovein": dtmovein_val.isoformat()
                if pd.notna(dtmovein_val)
                else None,
                "dtmoveout": dtmoveout_val.isoformat()
                if pd.notna(dtmoveout_val)
                else None,
                # 0–100 eviction risk score from transaction model
                "eviction_risk_score": float(row.get("eviction_risk_score") or 0.0),
                # Per-tenant payment / collections metrics (for property analytics)
                "dnumnsf": _safe_int_value(row.get("dnumnsf")),
                "dnumlate": _safe_int_value(row.get("dnumlate")),
                "damoutcollections": _safe_float_value(row.get("damoutcollections")),
                "drentwrittenoff": _safe_float_value(row.get("drentwrittenoff")),
                "dnonrentwrittenoff": _safe_float_value(row.get("dnonrentwrittenoff")),
                # New raw fields for potential UI use/debug
                "daypaid": _safe_int_value(row.get("daypaid")),
                "dpaysourcechange": _safe_int_value(row.get("dpaysourcechange")),
                "spaymentsource": row.get("spaymentsource"),
                # Per-tenant top drivers (now can include daypaid/dpaysourcechange/spaymentsource)
                "drivers": top_drivers,
            }
        )

    return out



def _compute_screening_model_payload():
    """
    Train a CatBoost model using **screening-only features** (plus sevicted
    label from transacts) to predict eviction (sevicted), then score the
    2024+ cohort that has screening data.

    In addition to per-tenant eviction_risk_score, this function also
    computes the top 3 driver features (with comparison to a low-risk
    baseline) so the frontend at-risk view can show local explanations.
    """
    import pandas as pd
    import numpy as np
    from sklearn.model_selection import train_test_split
    from catboost import CatBoostClassifier, Pool
    from datetime import date, datetime as dt

    # ------------------------------------------------------------------
    # Helper functions (adapted from NEW MODEL)
    # ------------------------------------------------------------------
    def clean_binary_flag(series: pd.Series) -> pd.Series:
        """
        Normalize common binary encodings to {0,1}.
        Handles:
        - booleans
        - 0/1
        - 'Y'/'N', 'YES'/'NO'
        - 'TRUE'/'FALSE'
        """
        s = series.copy()

        # If already numeric-ish, coerce and return
        if pd.api.types.is_numeric_dtype(s):
            return pd.to_numeric(s, errors="coerce")

        s = s.astype(str).str.strip().str.upper()
        mapping = {
            "1": 1,
            "0": 0,
            "Y": 1,
            "N": 0,
            "YES": 1,
            "NO": 0,
            "TRUE": 1,
            "FALSE": 0,
        }
        s = s.map(mapping)
        return s

    def coerce_numeric(series: pd.Series) -> pd.Series:
        """Coerce numeric-like strings (optionally with %) to float."""
        s = series.astype(str).str.replace("%", "", regex=False)
        return pd.to_numeric(s, errors="coerce")

    def combine_years_months(df: pd.DataFrame, years_col: str, months_col: str) -> pd.Series:
        """Combine years + months into total months."""
        years = df[years_col] if years_col in df.columns else 0
        months = df[months_col] if months_col in df.columns else 0
        years = pd.to_numeric(years, errors="coerce").fillna(0)
        months = pd.to_numeric(months, errors="coerce").fillna(0)
        return years * 12 + months

    def _safe_float(v):
        return float(v) if v is not None and not pd.isna(v) else None

    # Map DB column names -> canonical NEW MODEL names
    RENAME_MAP = {
        # dates
        "appcreddate": "applicant_credit_date",
        # booleans / flags
        "creditrun": "credit_run",
        "hascpmess": "has_checkpoint_msgs",
        "hasconsstmt": "has_consumer_stmt",
        # employment / residence tenure
        "currempmon": "current_emp_months",
        "currempyear": "current_emp_years",
        "currresmon": "current_res_months",
        "currresyear": "current_res_years",
        "prevempmon": "previous_emp_months",
        "prevempyear": "previous_emp_years",
        "prevresmon": "previous_res_months",
        "prevresyear": "previous_res_years",
        # income / debt
        "primincome": "primary_income",
        "addincome": "additional_income",
        "riskscore": "risk_score",
        "rentincratio": "rent_to_income_ratio_pct",
        "debtincratio": "debt_to_income_ratio_pct",
        "debtcredratio": "debt_to_credit_ratio_pct",
        "studdebt": "student_debt",
        "meddebt": "medical_debt",
        "totscordebt": "total_scorable_debt",
        "totdebt": "total_debt",
        "appmoninc": "application_monthly_income",
        "apptotdebt": "application_total_debt_policy",
        "avgriskscore": "avg_risk_score",
        # ids / names
        "voyappcode": "voyager_applicant_code",
        "voypropcode": "voyager_property_code",
        "propertyid": "property_id",
        "companyname": "company_name",
        "companycode": "company_code",
        "propname": "property_name",
        "voypropname": "voyager_property_name",
        "appstatus": "applicant_status",
        "scoremodel": "score_model",
        # free-text reason / checkpoint / review
        "reasonone": "reason_1",
        "reasontwo": "reason_2",
        "reasonthree": "reason_3",
        "checkmes1": "checkpoint_message_1",
        "checkmes2": "checkpoint_message_2",
        "itemrev1": "item_to_review_1",
        "itemrev2": "item_to_review_2",
        "itemrev3": "item_to_review_3",
    }

    # Which raw screening fields we want to expose as "driver" candidates
    # Which raw screening fields we want to expose as "driver" candidates.
    # These are also used as features in the screening model.
    # `direction` tells _compute_top_drivers whether higher or lower
    # than the low-risk baseline is considered worse.
    SCREENING_DRIVER_SPECS = {
        "riskscore": {
            "label": "Screening risk score",
            "direction": "low",  # lower score = worse (e.g., 0 vs 719)
        },
        "rentincratio": {
            "label": "Rent-to-income (%)",
            "direction": "high",  # higher ratio = worse
        },
        "debtincratio": {
            "label": "Debt-to-income (%)",
            "direction": "high",  # higher ratio = worse
        },
        "totdebt": {
            "label": "Total debt ($)",
            "direction": "high",  # more debt = worse
        },
    }

    def _prepare_screening_features(
        df_raw: pd.DataFrame,
        is_train: bool,
        trained_feature_cols=None,
        trained_categorical_cols=None,
    ):
        """
        Shared feature pipeline for training & scoring, adapted from NEW MODEL.

        When is_train=True: returns (X, y, feature_cols, categorical_feature_cols)
        When is_train=False: returns (X, None, feature_cols, categorical_feature_cols) but
        uses `trained_feature_cols` / `trained_categorical_cols` to align columns.
        """
        df = df_raw.copy()

        # 0) Rename DB columns -> canonical names used in NEW MODEL
        df.rename(columns=RENAME_MAP, inplace=True)
        df.columns = [str(c).strip() for c in df.columns]

        target_col = "sevicted" if is_train else None

        # --- Handle label for training ---
        if is_train:
            if target_col not in df.columns:
                return None, None, None, None

            df = df[df[target_col].notna()].copy()
            df[target_col] = clean_binary_flag(df[target_col])
            df = df[df[target_col].isin([0, 1])].copy()
            if df[target_col].nunique() < 2:
                return None, None, None, None

        # --- Type conversions: dates, numerics, booleans ---
        date_cols_expected = ["applicant_credit_date", "date"]
        date_cols = [c for c in date_cols_expected if c in df.columns]
        for c in date_cols:
            df[c] = pd.to_datetime(cast(df[c], "datetime64[ns]"), errors="coerce") if False else pd.to_datetime(df[c], errors="coerce")

        numeric_cols_expected = [
            "age",
            "current_emp_months",
            "current_emp_years",
            "current_res_months",
            "current_res_years",
            "previous_emp_months",
            "previous_emp_years",
            "previous_res_months",
            "previous_res_years",
            "income",
            "primary_income",
            "additional_income",
            "risk_score",
            "rent",
            "rent_to_income_ratio_pct",
            "debt_to_income_ratio_pct",
            "debt_to_credit_ratio_pct",
            "student_debt",
            "medical_debt",
            "total_scorable_debt",
            "total_debt",
            "application_monthly_income",
            "application_total_debt_policy",
            "avg_risk_score",
        ]
        numeric_cols = [c for c in numeric_cols_expected if c in df.columns]
        for c in numeric_cols:
            df[c] = coerce_numeric(df[c])

        bool_like_cols_expected = [
            "credit_run",
            "has_checkpoint_msgs",
            "has_consumer_stmt",
        ]
        for c in bool_like_cols_expected:
            if c in df.columns:
                df[c] = clean_binary_flag(df[c]).astype("float")

        # --- Feature engineering ---
        # Tenure in months
        if "current_emp_years" in df.columns or "current_emp_months" in df.columns:
            df["current_emp_tenure_months"] = combine_years_months(
                df, "current_emp_years", "current_emp_months"
            )

        if "current_res_years" in df.columns or "current_res_months" in df.columns:
            df["current_res_tenure_months"] = combine_years_months(
                df, "current_res_years", "current_res_months"
            )

        if "previous_emp_years" in df.columns or "previous_emp_months" in df.columns:
            df["previous_emp_tenure_months"] = combine_years_months(
                df, "previous_emp_years", "previous_emp_months"
            )

        if "previous_res_years" in df.columns or "previous_res_months" in df.columns:
            df["previous_res_tenure_months"] = combine_years_months(
                df, "previous_res_years", "previous_res_months"
            )

        # Normalize percentage ratios to 0–1
        ratio_pct_cols = [
            "rent_to_income_ratio_pct",
            "debt_to_income_ratio_pct",
            "debt_to_credit_ratio_pct",
        ]
        for c in ratio_pct_cols:
            if c in df.columns:
                df[c.replace("_pct", "_ratio")] = df[c] / 100.0

        # Flags for having student / medical debt
        if "student_debt" in df.columns:
            df["has_student_debt"] = (df["student_debt"].fillna(0) > 0).astype(int)

        if "medical_debt" in df.columns:
            df["has_medical_debt"] = (df["medical_debt"].fillna(0) > 0).astype(int)

        # Primary income share
        if "primary_income" in df.columns and "income" in df.columns:
            df["primary_income_share"] = np.where(
                (df["income"] > 0) & df["income"].notna(),
                df["primary_income"] / df["income"],
                np.nan,
            )

        # Log transforms for skewed amounts
        log_cols = [
            "income",
            "application_monthly_income",
            "total_debt",
            "total_scorable_debt",
            "student_debt",
            "medical_debt",
            "rent",
        ]
        for c in log_cols:
            if c in df.columns:
                df[f"log_{c}"] = np.log1p(df[c].clip(lower=0))

        # Date parts from screening date
        date_col_for_features = None
        if "applicant_credit_date" in df.columns:
            date_col_for_features = "applicant_credit_date"
        elif "date" in df.columns:
            date_col_for_features = "date"

        if date_col_for_features is not None:
            df[f"{date_col_for_features}_year"] = df[date_col_for_features].dt.year
            df[f"{date_col_for_features}_month"] = df[date_col_for_features].dt.month
            df[f"{date_col_for_features}_dayofweek"] = df[
                date_col_for_features
            ].dt.dayofweek

        # If we're *only* scoring, align to trained feature set and bail early
        if not is_train:
            if trained_feature_cols is None:
                return None, None, None, None

            for c in trained_feature_cols:
                if c not in df.columns:
                    df[c] = np.nan

            X = df[trained_feature_cols].copy()
            return X, None, trained_feature_cols, trained_categorical_cols

        # ------------------------------------------------------------------
        # Training-time feature selection / leakage control (NEW MODEL logic)
        # ------------------------------------------------------------------
        id_cols = [
            "applicant_credit_applicant_id",
            "applicant_credit_id",
            "applicant_id",
            "voyager_applicant_code",
            "voyager_property_code",
            "property_id",
            "application_id",
        ]
        name_cols = [
            "company_name",
            "company_code",
            "property_name",
            "voyager_property_name",
        ]
        free_text_cols = [
            "reason_1",
            "reason_2",
            "reason_3",
            "checkpoint_message_1",
            "checkpoint_message_2",
            "item_to_review_1",
            "item_to_review_2",
            "item_to_review_3",
        ]
        leakage_cols = ["applicant_status"]

        cols_to_exclude = set()
        for c in id_cols + name_cols + free_text_cols + leakage_cols + [target_col]:
            if c in df.columns:
                cols_to_exclude.add(c)
        for c in date_cols:
            if c in df.columns:
                cols_to_exclude.add(c)

        feature_cols = [c for c in df.columns if c not in cols_to_exclude]

        numeric_feature_candidates = [
            c for c in feature_cols if pd.api.types.is_numeric_dtype(df[c])
        ]
        categorical_feature_candidates = [
            c for c in feature_cols if not pd.api.types.is_numeric_dtype(df[c])
        ]

        # Force some numeric-looking cols to categorical
        force_categorical = [c for c in ["category", "score_model", "zip"]
                             if c in numeric_feature_candidates]
        numeric_feature_candidates = [
            c for c in numeric_feature_candidates if c not in force_categorical
        ]
        categorical_feature_candidates = (
            categorical_feature_candidates + force_categorical
        )

        # Drop redundant numeric inputs (keep derived features instead)
        optional_drop_numeric = [
            "current_emp_months",
            "current_emp_years",
            "current_res_months",
            "current_res_years",
            "previous_emp_months",
            "previous_emp_years",
            "previous_res_months",
            "previous_res_years",
            "rent_to_income_ratio_pct",
            "debt_to_income_ratio_pct",
            "debt_to_credit_ratio_pct",
        ]
        numeric_feature_candidates = [
            c for c in numeric_feature_candidates if c not in optional_drop_numeric
        ]

        feature_cols = numeric_feature_candidates + categorical_feature_candidates

        X = df[feature_cols].copy()
        y = df[target_col].astype(int).copy()

        # Drop degenerate features (all-missing or single level)
        all_missing_cols = [c for c in X.columns if X[c].isna().all()]
        single_level_cols = [
            c for c in X.columns if X[c].dropna().nunique() <= 1
        ]
        drop_cols = sorted(set(all_missing_cols + single_level_cols))
        if drop_cols:
            X = X.drop(columns=drop_cols)

        feature_cols = X.columns.tolist()
        categorical_feature_candidates = [
            c for c in feature_cols if not pd.api.types.is_numeric_dtype(X[c])
        ]

        return X, y, feature_cols, categorical_feature_candidates

    # ------------------------------------------------------------------
    # 1. Build training set: screening rows with known sevicted label
    # ------------------------------------------------------------------
    q_train = """
        SELECT
            s.*,
            t.sevicted
        FROM screening s
        INNER JOIN transacts t
            ON t.tscode = s.voyappcode
        WHERE t.sevicted IS NOT NULL;
    """
    train_df_raw = pd.read_sql(q_train, conn)

    if train_df_raw.empty:
        return {}

    # Baseline & spread for screening driver features (low-risk subset)
    baseline_screen = {}
    spread_screen = {}
    try:
        sev_series = clean_binary_flag(train_df_raw["sevicted"])
        low_risk_train = train_df_raw[sev_series == 0].copy()
        if not low_risk_train.empty:
            for col in SCREENING_DRIVER_SPECS.keys():
                if col not in low_risk_train.columns:
                    continue
                col_series = pd.to_numeric(low_risk_train[col], errors="coerce")
                col_series = col_series.dropna()
                if col_series.empty:
                    continue
                baseline_screen[col] = float(col_series.median())
                std_val = float(col_series.std(ddof=0))
                spread_screen[col] = std_val if std_val > 0 else None
    except Exception:
        baseline_screen = {}
        spread_screen = {}

    X_all, y_all, feature_cols, cat_cols = _prepare_screening_features(
        train_df_raw, is_train=True
    )

    if X_all is None or y_all is None:
        return {}
    if len(X_all) < 100 or y_all.nunique() < 2:
        # Not enough labelled data for a stable model
        return {}

    # ------------------------------------------------------------------
    # 2. Train/val/test split (60/20/20) and CatBoost model
    # ------------------------------------------------------------------
    X_train_full, X_test, y_train_full, y_test = train_test_split(
        X_all, y_all, test_size=0.2, random_state=42, stratify=y_all
    )
    X_train, X_val, y_train, y_val = train_test_split(
        X_train_full,
        y_train_full,
        test_size=0.25,
        random_state=42,
        stratify=y_train_full,
    )  # 0.25 of 0.8 => 0.2 => 60/20/20

    def _prep_catboost_frames(X: pd.DataFrame):
        X_cb = X.copy()
        for c in cat_cols or []:
            if c in X_cb.columns:
                X_cb[c] = X_cb[c].astype("string").fillna("MISSING")
        cat_idx = [
            X_cb.columns.get_loc(c)
            for c in (cat_cols or [])
            if c in X_cb.columns
        ]
        return X_cb, cat_idx

    X_train_cb, cat_idx = _prep_catboost_frames(X_train)
    X_val_cb, _ = _prep_catboost_frames(X_val)
    X_test_cb, _ = _prep_catboost_frames(X_test)

    cb_model = CatBoostClassifier(
        loss_function="Logloss",
        eval_metric="AUC",
        depth=4,
        learning_rate=0.06,
        l2_leaf_reg=1,
        iterations=1000,
        random_state=42,
        auto_class_weights="Balanced",
        verbose=False,
        early_stopping_rounds=50,
    )

    cb_model.fit(
        X_train_cb,
        y_train,
        cat_features=cat_idx,
        eval_set=(X_val_cb, y_val),
        use_best_model=True,
    )

    # ------------------------------------------------------------------
    # Global top drivers for the screening model (feature importance)
    # ------------------------------------------------------------------
    try:
        from catboost import Pool as CBPool

        train_pool = CBPool(X_train_cb, label=y_train, cat_features=cat_idx)
        importances = cb_model.get_feature_importance(train_pool)
        feat_names = list(X_train_cb.columns)

        pairs = list(zip(feat_names, importances))
        pairs.sort(key=lambda x: x[1], reverse=True)

        global_screen_drivers = []
        for feat_name, imp in pairs[:3]:
            global_screen_drivers.append(
                {
                    "feature_key": feat_name,
                    "feature_label": _pretty_screening_feature_label(feat_name),
                    "importance": float(imp),
                }
            )

        _SCREENING_MODEL_CACHE["global_drivers"] = global_screen_drivers
    except Exception as e:
        print(f"Could not compute global screening drivers: {e}")
        _SCREENING_MODEL_CACHE["global_drivers"] = []

    # ------------------------------------------------------------------
    # 3. Scoring cohort: 2024+ tenants with screening rows
    # ------------------------------------------------------------------
    q_score = """
        SELECT
            t.pscode,
            t.tscode,
            t.uscode,
            t.dtmovein,
            t.dtmoveout,
            t.dnumnsf,
            t.dnumlate,
            t.damoutcollections,
            t.drentwrittenoff,
            t.dnonrentwrittenoff,
            s.*
        FROM transacts t
        INNER JOIN screening s
            ON t.tscode = s.voyappcode
        WHERE t.pscode IS NOT NULL
          AND t.tscode IS NOT NULL
          AND t.dtmovein IS NOT NULL
          AND t.dtmovein >= DATE '2024-01-01';
    """

    score_df_raw = pd.read_sql(q_score, conn)

    if score_df_raw.empty:
        return {}

    # Keep only rows with *some* screening numeric info present,
    # so the screening view only shows tenants that truly have screening data.
    key_screen_cols = [
        c for c in ["riskscore", "totdebt", "rentincratio", "debtincratio"]
        if c in score_df_raw.columns
    ]
    if key_screen_cols:
        mask_has_screen = score_df_raw[key_screen_cols].notna().any(axis=1)
        score_df_raw = score_df_raw[mask_has_screen].copy()
        if score_df_raw.empty:
            return {}

    # Meta columns for output (use DB names here)
    meta_cols = [
        "pscode",
        "tscode",
        "uscode",
        "dtmovein",
        "dtmoveout",
        "dnumnsf",
        "dnumlate",
        "damoutcollections",
        "drentwrittenoff",
        "dnonrentwrittenoff",
        "riskscore",
        "totdebt",
        "rentincratio",
        "debtincratio",
    ]

    meta_cols = [c for c in meta_cols if c in score_df_raw.columns]
    meta_df = score_df_raw[meta_cols].copy()

    # Build feature matrix using *same* feature_cols / cat_cols as training
    X_score, _, _, _ = _prepare_screening_features(
        score_df_raw,
        is_train=False,
        trained_feature_cols=feature_cols,
        trained_categorical_cols=cat_cols,
    )
    if X_score is None or X_score.empty:
        return {}

    X_score_cb, _ = _prep_catboost_frames(X_score)

    # ------------------------------------------------------------------
    # 4. Predict probabilities and convert to 0–100 eviction risk scores
    # ------------------------------------------------------------------
    proba = cb_model.predict_proba(X_score_cb)[:, 1]
    scores_0_100 = (proba * 100.0).round(1)

    # ------------------------------------------------------------------
    # 5. Build property -> tenants mapping payload
    # ------------------------------------------------------------------
    out = {}

    for i, (_, row) in enumerate(meta_df.iterrows()):
        pscode = row.get("pscode")
        if not pscode:
            continue

        dtmovein_val = row.get("dtmovein")
        dtmoveout_val = row.get("dtmoveout")

        # Normalize dates to ISO
        def _to_iso(d):
            if d is None or pd.isna(d):
                return None
            if isinstance(d, pd.Timestamp):
                return d.date().isoformat()
            if isinstance(d, date):
                return d.isoformat()
            try:
                return pd.to_datetime(d).date().isoformat()
            except Exception:
                return None

        # Per-tenant top drivers for screening model (based on raw screening cols)
        top_drivers = _compute_top_drivers(
            row,
            driver_specs=SCREENING_DRIVER_SPECS,
            baseline=baseline_screen,
            spread=spread_screen,
            max_drivers=3,
        )

        tenant_entry = {
            "tscode": row.get("tscode"),
            "uscode": row.get("uscode"),
            "dtmovein": _to_iso(dtmovein_val),
            "dtmoveout": _to_iso(dtmoveout_val),
            # Screening-table fields
            "riskscore": _safe_float(row.get("riskscore")),
            "totdebt": _safe_float(row.get("totdebt")),
            "rentincratio": _safe_float(row.get("rentincratio")),
            "debtincratio": _safe_float(row.get("debtincratio")),
            # Joined-in payment / collections metrics from transacts
            "dnumnsf": _safe_int_value(row.get("dnumnsf")),
            "dnumlate": _safe_int_value(row.get("dnumlate")),
            "damoutcollections": _safe_float_value(row.get("damoutcollections")),
            "drentwrittenoff": _safe_float_value(row.get("drentwrittenoff")),
            "dnonrentwrittenoff": _safe_float_value(row.get("dnonrentwrittenoff")),
            # Screening-model eviction risk, scaled 0–100
            "eviction_risk_score": float(scores_0_100[i]),
            # Per-tenant top three driver features for at-risk view
            "drivers": top_drivers,
        }

        out.setdefault(pscode, []).append(tenant_entry)

    return out


def _compute_top_drivers(row, driver_specs, baseline, spread, max_drivers=3):
    """
    Generic helper to compute the top-N driver features for a single tenant.

    Parameters
    ----------
    row : pandas.Series
        Row containing all raw/engineered feature columns.
    driver_specs : dict
        Either {feature_name: "Label"} or
        {feature_name: {"label": ..., "direction": ...}} where
        direction ∈ {"high", "low", "distance", "category"}.
    baseline : dict
        Mapping feature_name -> baseline (e.g., median or modal value for
        low-risk tenants).
    spread : dict
        Mapping feature_name -> scale (std-dev) used to normalize differences
        (ignored for "category" direction).
    max_drivers : int
        Maximum number of drivers to return.

    Returns
    -------
    list[dict]
        Each dict has: feature_key, feature_label, value, baseline, impact_score.
        Only features where the tenant is *worse than* the low-risk baseline
        (according to direction) are returned.
    """
    drivers = []

    for feature_key, spec in driver_specs.items():
        if feature_key not in row.index:
            continue

        # Interpret spec
        if isinstance(spec, str):
            label = spec
            direction = "distance"
        elif isinstance(spec, dict):
            label = spec.get("label", feature_key)
            direction = spec.get("direction", "distance")
        else:
            label = str(spec)
            direction = "distance"

        val = row.get(feature_key)
        base = baseline.get(feature_key)

        if val is None or base is None:
            continue
        if pd.isna(val) or pd.isna(base):
            continue

        # --- Categorical case: compare category vs baseline category
        if direction == "category":
            val_str = str(val)
            base_str = str(base)
            # If same as low-risk baseline, not a "driver"
            if val_str == base_str:
                continue

            impact = 1.0  # treat any deviation from baseline as a unit impact
            drivers.append(
                {
                    "feature_key": feature_key,
                    "feature_label": label,
                    "value": val_str,
                    "baseline": base_str,
                    "impact_score": float(impact),
                }
            )
            continue

        # --- Numeric path (existing behavior, with "high"/"low"/"distance") ---
        try:
            v_float = float(val)
            b_float = float(base)
        except (TypeError, ValueError):
            continue

        diff = v_float - b_float
        scale = spread.get(feature_key)

        # Normalize by spread if available
        if scale is None or not np.isfinite(scale) or scale == 0:
            norm_diff = diff
        else:
            norm_diff = diff / scale

        # Orientation: only keep features where the tenant is "worse"
        if direction == "high":
            # higher than baseline is worse; ignore better-than-baseline
            if norm_diff <= 0:
                continue
            impact = norm_diff
        elif direction == "low":
            # lower than baseline is worse
            if norm_diff >= 0:
                continue
            impact = -norm_diff
        else:  # "distance" – fall back to absolute difference
            impact = abs(norm_diff)

        if not np.isfinite(impact) or impact <= 0:
            continue

        drivers.append(
            {
                "feature_key": feature_key,
                "feature_label": label,
                "value": float(v_float),
                "baseline": float(b_float),
                "impact_score": float(impact),
            }
        )

    drivers.sort(key=lambda d: d["impact_score"], reverse=True)
    return drivers[:max_drivers]


def _pretty_transaction_feature_label(name: str) -> str:
    """
    Human-readable labels for transaction-model features.
    """
    mapping = {
        "dnumnsf": "NSF count",
        "dnumlate": "Late payment count",
        "davgdayslate": "Average days late",
        "srent": "Monthly rent",
        "dincome": "Reported income",
        "late_ratio": "Late-payments / NSF ratio",
        "renewed_flag": "Renewed lease flag",
        "fulfilled_flag": "Fulfilled lease term flag",
        "rent_to_income": "Rent-to-income ratio",
        "tenure_days": "Tenure length (days)",
        "daypaid": "Day-of-month paid",
        "dpaysourcechange": "Payment-source changes",
        "spaymentsource": "Payment source",
    }
    if name in mapping:
        return mapping[name]

    # Fallback: unsnake
    label = name.replace("_", " ").strip()
    if not label:
        return name
    return label[0].upper() + label[1:]



def _pretty_screening_feature_label(raw: str) -> str:
    """
    Human-readable labels for screening-model features.
    If it's a log_ feature (log_xxx), strip 'log_' so we show the
    underlying variable, not the normalized/logged one.
    """
    if not raw:
        return raw

    base = raw
    if base.startswith("log_"):
        base = base[4:]   # strip 'log_'

    mapping = {
        "risk_score": "Screening risk score",
        "rent_to_income_ratio": "Rent-to-income ratio",
        "debt_to_income_ratio": "Debt-to-income ratio",
        "debt_to_credit_ratio": "Debt-to-credit ratio",
        "income": "Reported income",
        "total_debt": "Total debt",
        "total_scorable_debt": "Total scorable debt",
        "application_monthly_income": "Application monthly income",
        "application_total_debt_policy": "Application total debt (policy)",
        "avg_risk_score": "Average risk score",
        "primary_income_share": "Primary income share",
        "current_emp_tenure_months": "Current employment tenure (months)",
        "current_res_tenure_months": "Current residence tenure (months)",
        "previous_emp_tenure_months": "Previous employment tenure (months)",
        "previous_res_tenure_months": "Previous residence tenure (months)",
        "has_student_debt": "Has student debt flag",
        "has_medical_debt": "Has medical debt flag",
    }

    if base in mapping:
        return mapping[base]

    # Last-resort fallback: unsnake and capitalize
    label = base.replace("_", " ").strip()
    if not label:
        label = raw
    return label[0].upper() + label[1:]

def _safe_int_value(v):
    """
    Convert a value to a plain Python int for JSON (NaN/None -> 0).
    """
    try:
        if v is None or (isinstance(v, (float, np.floating)) and np.isnan(v)):
            return 0
    except Exception:
        pass
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def _safe_float_value(v):
    """
    Convert a value to a plain Python float for JSON (NaN/None -> 0.0).
    """
    try:
        if v is None or (isinstance(v, (float, np.floating)) and np.isnan(v)):
            return 0.0
    except Exception:
        pass
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# @app.route("/upload", methods=["POST"])
# def upload_file():
#     file = None
#     dataName = None
#     if "transact" in request.files:
#         file = request.files["transact"]
#         dataName = "transacts"
#         PRIMARY_KEY = "tscode"
#     elif "screening" in request.files:
#         file = request.files["screening"]
#         dataName = "screening"
#         PRIMARY_KEY = "voyAppCode"
#         col_map = COLUMN_MAPS.get(dataName, {})

#         def map_columns(row):
#             mapped = {}
#             for key, value in row.items():
#                 mapped[col_map.get(key.strip(), key)] = value
#             return mapped

#     else:
#         return jsonify({"error": "No file uploaded"}), 400
    
#     filename = file.filename
#     content = file.read()  # bytes

#     ext = os.path.splitext(filename)[1].lower()
#     if ext == ".csv":
#         try:
#             csv_file = io.StringIO(content.decode("utf-8-sig"))
#         except UnicodeDecodeError:
#             return jsonify({"message": f"{filename} is not UTF-8 encoded"}), 400

#         # Skip first 5 rows (headers/noise)
#         # for _ in range(5):
#         #     next(csv_file, None)

#         reader = csv.DictReader(csv_file)
#         reader.fieldnames = [h.strip().lower() for h in reader.fieldnames]

#         # Prepare upsert dynamically based on first row
#         first = next(reader, None)
#         if first is None:
#             return jsonify({"message": "No rows detected after header"}), 400

#         first = _normalize_row(first)
#         if dataName == "screening":
#             first = map_columns(first)   # apply COLUMN_MAPS here
#         columns = list(first.keys())
#         placeholders = ', '.join(['%s'] * len(columns))
#         colsql = ', '.join(columns)
#         update_clause = ', '.join([f"{c}=EXCLUDED.{c}" for c in columns if c != PRIMARY_KEY])

#         sql = f"INSERT INTO {dataName} ({colsql}) VALUES ({placeholders}) ON CONFLICT ({PRIMARY_KEY}) DO UPDATE SET {update_clause}"

#         batch, batch_size = [], 1000

#         def add_row(r):
#             r = _normalize_row(r)
#             if dataName == "screening":
#                 r = map_columns(r)
#             if not r.get(PRIMARY_KEY):
#                 return
#             batch.append([r.get(c) for c in columns])

#         add_row(first)
#         for r in reader:
#             add_row(r)
#             if len(batch) >= batch_size:
#                 cursor.executemany(sql, batch)
#                 batch.clear()
#         if batch:
#             cursor.executemany(sql, batch)

#     elif ext == ".xlsx":
#         xlsx = io.BytesIO(content)
#         wb = load_workbook(xlsx, data_only=True)
#         ws = wb.active
#         rows = list(ws.iter_rows(values_only=True))
#         rows = rows[5:]  # skip first 5 rows
#         if not rows:
#             return jsonify({"message": "No data rows detected"}), 400
#         headers = [str(h).strip().lower() for h in rows[0]]
#         for row in rows[1:]:
#             rd = {k: v for k, v in zip(headers, row)}
#             rd = _normalize_row(rd)
#             if dataName == "screening":
#                 rd = map_columns(rd)
#             if not rd.get(PRIMARY_KEY):
#                 continue
#             cols = list(rd.keys())
#             placeholders = ', '.join(['%s'] * len(cols))
#             colsql = ', '.join(cols)
#             update_clause = ', '.join([f"{c}=EXCLUDED.{c}" for c in cols if c != PRIMARY_KEY])
#             sql = f"INSERT INTO transacts ({colsql}) VALUES ({placeholders}) ON CONFLICT ({PRIMARY_KEY}) DO UPDATE SET {update_clause}"
#             cursor.execute(sql, [rd.get(c) for c in cols])
#     else:
#         return jsonify({"message": "Unsupported file type"}), 400

#     # Touch meta_updates
#     cursor.execute("INSERT INTO meta_updates (updated_at) VALUES (NOW())")
#     return jsonify({"message": f"{filename} uploaded successfully"}), 200

@app.route("/tenants/screening-eviction-risk", methods=["GET"])
def get_tenants_screening_eviction_risk():
    """
    Screening+transactions model eviction risk scores (0–100) for the same
    active cohort as /tenants/active, grouped by property code.

    Shape:
      {
        pscode: [
          {
            "tscode": ...,
            "uscode": ...,
            "dtmovein": "YYYY-MM-DD",
            "dtmoveout": "YYYY-MM-DD" | None,
            "riskscore": float | None,
            "totdebt": float | None,
            "rentincratio": float | None,
            "debtincratio": float | None,
            "eviction_risk_score": float | None
          },
          ...
        ]
      }
    """
    try:
        current_ts = _latest_meta_ts()

        if (
            _SCREENING_MODEL_CACHE.get("payload") is not None
            and _SCREENING_MODEL_CACHE.get("last_meta_ts") == current_ts
        ):
            return jsonify(_SCREENING_MODEL_CACHE["payload"]), 200

        payload = _compute_screening_model_payload()
        _SCREENING_MODEL_CACHE["payload"] = payload
        _SCREENING_MODEL_CACHE["last_meta_ts"] = current_ts

        return jsonify(payload), 200
    except Exception as e:
        print(f"Screening eviction risk model error: {str(e)}")
        import traceback

        traceback.print_exc()
        # Fail soft – frontend will just see no tenants for this view.
        return jsonify({}), 200


@app.route("/features/importance", methods=["GET", "OPTIONS"])
def feature_importance():
    if request.method == "OPTIONS":
        return ("", 204)

    try:
        current_ts = _latest_meta_ts()

        # If cache is valid, serve instantly
        if (
            _FEATURE_IMPORTANCE_CACHE["payload"] is not None
            and _FEATURE_IMPORTANCE_CACHE["last_meta_ts"] == current_ts
        ):
            return jsonify(_FEATURE_IMPORTANCE_CACHE["payload"]), 200

        # Otherwise compute, store, return
        payload = _compute_feature_importance_payload()
        _FEATURE_IMPORTANCE_CACHE["payload"] = payload
        _FEATURE_IMPORTANCE_CACHE["last_meta_ts"] = current_ts

        return jsonify(payload), 200

    except Exception as e:
        print(f"Feature importance error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "auc": None,
            "top_features": []
        }), 200
    
@app.route("/tenants/eviction-risk", methods=["GET"])
def get_tenants_eviction_risk():
    """
    Transaction-model eviction risk scores (0–100) for tenants whose
    lease start is in 2024 or later, grouped by property code.

    Shape mirrors /tenants/active: { pscode: [ { ... }, ... ] }.
    """
    try:
        current_ts = _latest_meta_ts()

        # Simple cache so we don't retrain the model for every property click
        if (
            _TRANSACTION_MODEL_CACHE.get("payload") is not None
            and _TRANSACTION_MODEL_CACHE.get("last_meta_ts") == current_ts
        ):
            return jsonify(_TRANSACTION_MODEL_CACHE["payload"]), 200

        payload = _compute_transaction_model_payload()
        _TRANSACTION_MODEL_CACHE["payload"] = payload
        _TRANSACTION_MODEL_CACHE["last_meta_ts"] = current_ts

        return jsonify(payload), 200
    except Exception as e:
        print(f"Eviction risk model error: {str(e)}")
        import traceback

        traceback.print_exc()
        # Fail soft – frontend will simply show "no tenants" for this view.
        return jsonify({}), 200


# -------------------------------------------------
# Fetch Active Tenants
# -------------------------------------------------

@app.route('/tenants/active', methods=['GET'])
@app.route('/tenants/active', methods=['GET'])
def get_tenants():
    query = """
    SELECT 
        t.pscode,
        t.tscode,
        t.uscode,
        t.dtmovein,
        t.dtmoveout,
        s.riskscore,
        s.totdebt,
        s.rentincratio,
        s.debtincratio
    FROM transacts t
    LEFT JOIN screening s
        ON t.tscode = s.voyappcode
    WHERE t.pscode IS NOT NULL
      AND t.tscode IS NOT NULL
      AND t.dtmovein IS NOT NULL
      AND t.dtmovein >= DATE '2024-01-01'
      AND (t.dtmoveout IS NULL OR t.dtmoveout > DATE '2025-04-01')
    ORDER BY t.dtmovein DESC, t.tscode;
    """

    try:
        cursor.execute(query)
        tenant_list = cursor.fetchall()

        tenant_mapping = {}
        for row in tenant_list:
            # unpack all fields returned by the query
            pscode, tscode, uscode, dtmovein, dtmoveout, riskscore, totdebt, rentincratio, debtincratio = row


            if pscode not in tenant_mapping:
                tenant_mapping[pscode] = []
            
            tenant_mapping[pscode].append({
                'tscode': tscode,
                'uscode': uscode,
                'dtmovein': dtmovein.isoformat() if dtmovein else None,
                'dtmoveout': dtmoveout.isoformat() if dtmoveout else None,
                'riskscore': riskscore,
                'totdebt': totdebt,
                'rentincratio': rentincratio,
                'debtincratio': debtincratio
            })

        return jsonify(tenant_mapping)
    
    except Exception as e:
        return jsonify({'Error': str(e)}), 500

@app.route("/models/global-drivers", methods=["GET"])
def models_global_drivers():
    """
    Return the overall top drivers of eviction risk for each model
    (screening + transactions), based on CatBoost feature importance.

    Shape:
      {
        "screening": {
          "top_drivers": [
            { "feature_key": ..., "feature_label": ..., "importance": float },
            ...
          ]
        },
        "transactions": {
          "top_drivers": [
            { "feature_key": ..., "feature_label": ..., "importance": float },
            ...
          ]
        }
      }
    """
    try:
        current_ts = _latest_meta_ts()

        # Warm transaction model cache if needed
        if (
            _TRANSACTION_MODEL_CACHE.get("payload") is None
            or _TRANSACTION_MODEL_CACHE.get("last_meta_ts") != current_ts
        ):
            tx_payload = _compute_transaction_model_payload()
            _TRANSACTION_MODEL_CACHE["payload"] = tx_payload
            _TRANSACTION_MODEL_CACHE["last_meta_ts"] = current_ts

        # Warm screening model cache if needed
        if (
            _SCREENING_MODEL_CACHE.get("payload") is None
            or _SCREENING_MODEL_CACHE.get("last_meta_ts") != current_ts
        ):
            sc_payload = _compute_screening_model_payload()
            _SCREENING_MODEL_CACHE["payload"] = sc_payload
            _SCREENING_MODEL_CACHE["last_meta_ts"] = current_ts

        screening_drivers = _SCREENING_MODEL_CACHE.get("global_drivers", []) or []
        tx_drivers = _TRANSACTION_MODEL_CACHE.get("global_drivers", []) or []

        return jsonify(
            {
                "screening": {"top_drivers": screening_drivers},
                "transactions": {"top_drivers": tx_drivers},
            }
        ), 200

    except Exception as e:
        print(f"Global drivers error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(
            {
                "screening": {"top_drivers": []},
                "transactions": {"top_drivers": []},
            }
        ), 200

# -------------------------------------------------
# Health
# -------------------------------------------------
@app.route("/health")
def health():
    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(port=5000, debug=True)
